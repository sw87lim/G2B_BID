"""
조달청 입찰공고 자동화 에이전트 - 최종 정리 버전
필터링 기준:
  - 공고일자: 오늘부터 7일 전까지
  - 또는 마감일자: 오늘부터 7일 이후까지
"""

import json
import os
import sys
import requests
import logging
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
import xml.etree.ElementTree as ET

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

SERVICE_KEY = "8d8ea48a831eff6b90dbcab10ecf0673aa56ba6bdeec8a24bcfc4bf820cadcad"
API_BASE = "http://apis.data.go.kr/1230000/ad/BidPublicInfoService"

ENDPOINTS = {
    "공사": "getBidPblancListInfoCnstwk",
    "물품": "getBidPblancListInfoBPay"
}

KEYWORDS = [
    "냉방", "난방", "에어컨", "GHP", "EHP", "기계", "에어콘",
    "공기조화", "실외기", "실내기", "히트펌프", "EHS", "P2H"
]

SMTP_SERVER = "smtp.gmail.com"
SMTP_PORT = 587
RECIPIENT_EMAIL = "sw87.lim@samsung.com"
G2B_BASE_URL = "https://www.g2b.go.kr"

def load_config():
    try:
        with open('config.json', 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        logger.error("❌ config.json을 찾을 수 없습니다!")
        sys.exit(1)

def matches_keywords(title):
    if not title:
        return False
    title_lower = title.lower()
    return any(kw.lower() in title_lower for kw in KEYWORDS)

def should_include_bid(bid):
    """
    필터링 기준:
    - 공고일자: 오늘부터 7일 전까지
    - 또는 마감일자: 오늘부터 7일 이후까지
    """
    try:
        today = datetime.now()
        
        공고일자_str = bid['공고일자']
        if len(공고일자_str) == 8:
            공고일자 = datetime.strptime(공고일자_str, "%Y%m%d")
        else:
            return False
        
        마감일자_str = bid['마감일자']
        if len(마감일자_str) == 8:
            마감일자 = datetime.strptime(마감일자_str, "%Y%m%d")
        else:
            return False
        
        # 기준1) 공고일자가 오늘부터 7일 전까지
        공고일자_범위_시작 = today - timedelta(days=7)
        기준1 = 공고일자_범위_시작.date() <= 공고일자.date() <= today.date()
        
        # 기준2) 마감일자가 오늘부터 7일 이후까지
        마감일자_범위_끝 = today + timedelta(days=7)
        기준2 = today.date() <= 마감일자.date() <= 마감일자_범위_끝.date()
        
        return 기준1 or 기준2
    
    except Exception as e:
        logger.warning(f"⚠️ 필터링 오류: {e}")
        return False

def fetch_bids(bid_type, start_date, end_date):
    endpoint = ENDPOINTS.get(bid_type)
    if not endpoint:
        return []
    
    all_bids = []
    page = 1
    
    logger.info(f"🔄 [{bid_type}] API 조회 중...")
    
    while True:
        try:
            url = f"{API_BASE}/{endpoint}"
            params = {
                'serviceKey': SERVICE_KEY,
                'pageNo': page,
                'numOfRows': 100,
                'inqryDiv': '1',
                'inqryBgnDt': start_date,
                'inqryEndDt': end_date
            }
            
            response = requests.get(url, params=params, timeout=30)
            
            if response.status_code != 200:
                break
            
            root = ET.fromstring(response.content)
            items = root.findall('.//item')
            
            if not items:
                break
            
            logger.info(f"  ✓ 페이지 {page}: {len(items)}건")
            
            for item in items:
                def get_text(tag, default=''):
                    val = item.findtext(tag)
                    return val if val else default
                
                bid = {
                    '공고번호': get_text('BidNO'),
                    '공고명': get_text('BidNm'),
                    '공고기관': get_text('InqryCorpNm'),
                    '입찰유형': bid_type,
                    '기초금액': int(get_text('BidPblancAmt', '0')) or 0,
                    '공고일자': get_text('BidPblancDt'),
                    '마감일자': get_text('BidClseDt'),
                    '입찰방식': get_text('BidMthdTpNm'),
                    '상태': get_text('BidPblancSt', '정상'),
                }
                all_bids.append(bid)
            
            page += 1
        
        except Exception as e:
            logger.error(f"  ❌ 오류: {str(e)}")
            break
    
    logger.info(f"✅ [{bid_type}] {len(all_bids)}건 조회 완료\n")
    return all_bids

def fetch_bid_details(bid_no):
    try:
        url = f"{API_BASE}/getBidPblancDetailInfoOpenAPI"
        params = {
            'serviceKey': SERVICE_KEY,
            'BidNO': bid_no
        }
        
        response = requests.get(url, params=params, timeout=15)
        
        if response.status_code != 200:
            return {}
        
        root = ET.fromstring(response.content)
        item = root.find('.//item')
        
        if item is None:
            return {}
        
        def get_text(tag, default=''):
            val = item.findtext(tag) if item else None
            return val if val else default
        
        return {
            '게시일시': get_text('BidPblancDt'),
            '개찰일자': get_text('BidOpnnDt'),
            '낙찰하한율': get_text('LowBidRate', '미지정'),
            '추정가격': int(get_text('BidEstmAmt', '0')) or 0,
            '지역제한': get_text('RgnRestrictYn', 'N'),
            '참가가능지역': get_text('RgnList', '전국'),
            '업종제한': get_text('IndstryRestrictYn', 'N'),
            '업종제한사항': get_text('IndstryRestrict', '없음'),
            '수요기관': get_text('RequiredCorpNm'),
            '납품기한': get_text('DeliveryDuedt'),
            '공고담당자': get_text('BidOfficerNm', ''),
            '부서명': get_text('DeptNm', ''),
            '담당자': get_text('ChargePerson', ''),
            '전화번호': get_text('PhoneNumber', ''),
        }
    
    except Exception as e:
        logger.warning(f"  ⚠️ 상세정보 조회 실패: {str(e)}")
        return {}

def calculate_remaining_days(deadline_str):
    try:
        if len(deadline_str) == 8:
            deadline = datetime.strptime(deadline_str, "%Y%m%d")
        else:
            deadline = datetime.strptime(deadline_str[:8], "%Y%m%d")
        
        remaining = (deadline - datetime.now()).days
        return remaining
    except:
        return -999

def generate_excel(filtered_bids):
    if not filtered_bids:
        logger.warning("❌ 필터링된 공고가 없습니다!")
        return None
    
    logger.info("📊 엑셀 파일 생성 중...")
    
    wb = Workbook()
    ws = wb.active
    ws.title = '입찰공고'
    
    headers = [
        'No.', '공고명', '게시일시', '개찰일자', '공고기관', '수요기관', '부서명',
        '입찰유형', '기초금액', '추정가격', '낙찰하한율', '마감일자', '남은일수',
        '참가가능지역', '업종제한사항', '공고담당자', '담당자', '전화번호', '상태'
    ]
    
    ws.append(headers)
    
    header_fill = PatternFill(start_color='2954D1', end_color='2954D1', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=11)
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    for row_idx, bid in enumerate(filtered_bids, 2):
        remaining_days = calculate_remaining_days(bid['마감일자'])
        
        def format_date(date_str):
            if date_str and len(date_str) == 8:
                return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
            return date_str
        
        row_data = [
            row_idx - 1,
            bid['공고명'],
            format_date(bid.get('게시일시', bid['공고일자'])),
            format_date(bid.get('개찰일자', '')),
            bid['공고기관'],
            bid.get('수요기관', ''),
            bid.get('부서명', ''),
            bid['입찰유형'],
            bid['기초금액'],
            bid.get('추정가격', 0),
            bid.get('낙찰하한율', '미지정'),
            format_date(bid['마감일자']),
            remaining_days,
            bid.get('참가가능지역', '전국'),
            bid.get('업종제한사항', '없음'),
            bid.get('공고담당자', ''),
            bid.get('담당자', ''),
            bid.get('전화번호', ''),
            bid['상태'],
        ]
        
        ws.append(row_data)
        
        for col_num in [9, 10]:
            ws.cell(row=row_idx, column=col_num).number_format = '#,##0'
    
    column_widths = [5, 15, 35, 12, 12, 15, 15, 8, 15, 15, 10, 12, 10, 15, 10, 10, 15, 12, 10]
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    
    ws.freeze_panes = 'B2'
    ws.auto_filter.ref = f'A1:S{len(filtered_bids) + 1}'
    
    filename = f"입찰공고_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    wb.save(filename)
    
    logger.info(f"✅ 엑셀 파일 생성: {filename}\n")
    return filename

def save_to_json(filtered_bids, filename='bid_data.json'):
    logger.info("💾 JSON 파일 생성 중...")
    
    data = []
    for idx, bid in enumerate(filtered_bids, 1):
        remaining_days = calculate_remaining_days(bid['마감일자'])
        
        def format_date(date_str):
            if date_str and len(date_str) == 8:
                return f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
            return date_str
        
        item = {
            '번호': idx,
            '입찰공고번호': bid['공고번호'],
            '공고명': bid['공고명'],
            '게시일시': format_date(bid.get('게시일시', bid['공고일자'])),
            '개찰일자': format_date(bid.get('개찰일자', '')),
            '공고기관': bid['공고기관'],
            '수요기관': bid.get('수요기관', ''),
            '부서명': bid.get('부서명', ''),
            '입찰유형': bid['입찰유형'],
            '기초금액': bid['기초금액'],
            '추정가격': bid.get('추정가격', 0),
            '낙찰하한율': bid.get('낙찰하한율', '미지정'),
            '마감일자': format_date(bid['마감일자']),
            '남은일수': remaining_days,
            '참가가능지역': bid.get('참가가능지역', '전국'),
            '업종제한사항': bid.get('업종제한사항', '없음'),
            '공고담당자': bid.get('공고담당자', ''),
            '담당자': bid.get('담당자', ''),
            '전화번호': bid.get('전화번호', ''),
            '상태': bid['상태'],
            '공고서_링크': G2B_BASE_URL,
        }
        data.append(item)
    
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        logger.info(f"✅ JSON 파일 생성: {filename}")
        logger.info(f"   {len(data)}건 저장됨\n")
        return filename
    except Exception as e:
        logger.error(f"❌ JSON 파일 생성 실패: {e}")
        return None

def send_email(filename, filtered_bids, config):
    if not config.get('smtp_user') or not config.get('smtp_password'):
        logger.warning("⚠️ 이메일 설정이 없습니다.")
        return False
    
    try:
        logger.info("📧 이메일 발송 중...")
        
        html_body = f"""
        <html>
        <head><meta charset="utf-8"></head>
        <body style="font-family: Arial; color: #333;">
            <h2>🏗️ 입찰공고 알림</h2>
            <p>{datetime.now().strftime('%Y년 %m월 %d일')}</p>
            <hr>
            <p><strong>📊 신규 공고: {len(filtered_bids)}건</strong></p>
            <p>첨부된 Excel 파일에서 상세 내용을 확인하세요.</p>
            <p><strong>필터링 기준:</strong></p>
            <ul>
                <li>✅ 공고일자: 오늘부터 7일 전까지</li>
                <li>또는</li>
                <li>✅ 마감일자: 오늘부터 7일 이후까지</li>
            </ul>
            <hr>
            <p style="color: #999; font-size: 12px;">자동 생성됨 | {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
        </body>
        </html>
        """
        
        msg = MIMEMultipart('alternative')
        msg['Subject'] = f"[입찰공고 알림] {datetime.now().strftime('%Y-%m-%d')} - {len(filtered_bids)}건"
        msg['From'] = config['smtp_user']
        msg['To'] = RECIPIENT_EMAIL
        
        msg.attach(MIMEText(html_body, 'html', 'utf-8'))
        
        if os.path.exists(filename):
            with open(filename, 'rb') as attachment:
                part = MIMEBase('application', 'octet-stream')
                part.set_payload(attachment.read())
                encoders.encode_base64(part)
                part.add_header('Content-Disposition', f'attachment; filename= {filename}')
                msg.attach(part)
        
        with smtplib.SMTP(SMTP_SERVER, SMTP_PORT) as server:
            server.starttls()
            server.login(config['smtp_user'], config['smtp_password'])
            server.sendmail(config['smtp_user'], RECIPIENT_EMAIL, msg.as_string())
        
        logger.info(f"✅ 이메일 발송 완료\n")
        return True
    
    except Exception as e:
        logger.error(f"❌ 이메일 발송 실패: {str(e)}")
        return False

def main():
    logger.info("=" * 60)
    logger.info("🚀 조달청 입찰공고 자동화 에이전트")
    logger.info("=" * 60 + "\n")
    
    config = load_config()
    
    today = datetime.now()
    end_date = today.strftime("%Y%m%d")
    start_date = (today - timedelta(days=7)).strftime("%Y%m%d")
    
    logger.info(f"📅 조회 범위: {start_date[:4]}-{start_date[4:6]}-{start_date[6:]} ~ {end_date[:4]}-{end_date[4:6]}-{end_date[6:]}\n")
    
    # API 조회
    all_bids = []
    for bid_type in ['공사', '물품']:
        bids = fetch_bids(bid_type, start_date, end_date)
        all_bids.extend(bids)
    
    if not all_bids:
        logger.error("❌ 조회된 공고가 없습니다!")
        return
    
    logger.info(f"📋 전체: {len(all_bids)}건\n")
    
    # 날짜 필터링
    logger.info("🔍 날짜 필터링 중...")
    filtered_by_date = [b for b in all_bids if should_include_bid(b)]
    logger.info(f"✅ {len(filtered_by_date)}건\n")
    
    # 키워드 필터링
    logger.info("🔍 키워드 필터링 중...")
    filtered_bids = [b for b in filtered_by_date if matches_keywords(b.get('공고명', ''))]
    
    if not filtered_bids:
        logger.error("❌ 조건을 만족하는 공고가 없습니다!")
        return
    
    logger.info(f"✅ {len(filtered_bids)}건\n")
    
    logger.info("📌 필터링된 공고:")
    for idx, bid in enumerate(filtered_bids, 1):
        remaining = calculate_remaining_days(bid['마감일자'])
        logger.info(f"  {idx}. [{bid['입찰유형']}] {bid['공고명'][:40]} (남은일수: {remaining}일)")
    logger.info()
    
    # 상세정보 조회
    logger.info("📊 상세정보 조회 중...\n")
    for bid in filtered_bids:
        details = fetch_bid_details(bid['공고번호'])
        bid.update(details)
    
    # 파일 생성
    excel_file = generate_excel(filtered_bids)
    json_file = save_to_json(filtered_bids)
    
    # 이메일 발송
    send_email(excel_file, filtered_bids, config)
    
    logger.info("=" * 60)
    logger.info("✅ 완료!")
    logger.info("=" * 60)

if __name__ == "__main__":
    main()
